"""
IIBA Spreadsheet Parser — Track A Ingest

Reads IIBA_Value_Streams_.xlsx and IIBA_Capability_Map.xlsx,
produces an IREngagementModel.

No LLM involvement. Deterministic.
"""

import re
import openpyxl
from pathlib import Path
from ir_types import (
    IREngagementModel, IRElement, SourceRecord, Confidence,
    ElementType, DiscoveredBy, ResolutionStatus,
)
from stage_names import STAGE_NAMES


# ── Normalisation ─────────────────────────────────────────────────

def normalise(s: str) -> str:
    """Collapse whitespace, newlines, nbsp."""
    return re.sub(r'\s+', ' ', s.replace('\n', ' ').replace('\xa0', ' ')).strip()


def make_id(prefix: str, name: str) -> str:
    """Deterministic ID from prefix + normalised name."""
    slug = re.sub(r'[^a-z0-9]+', '_', normalise(name).lower()).strip('_')
    return f"{prefix}_{slug}"


def derive_stage_name(description: str) -> str:
    """Derive a concise stage name from a narrative description."""
    desc = normalise(description)

    # Common patterns: "Verb X and Y; ..." or "Verb X through Y."
    # Take the first clause up to a semicolon, period, or dash
    first_clause = re.split(r'[;.–—]', desc)[0].strip()

    # If still long, take first N words
    words = first_clause.split()
    if len(words) > 8:
        first_clause = ' '.join(words[:7])

    # Capitalise first letter
    if first_clause:
        first_clause = first_clause[0].upper() + first_clause[1:]

    return first_clause


# ── Capability Map Parser ─────────────────────────────────────────

def parse_capability_map(filepath: str, source_id: str) -> list[IRElement]:
    """Parse IIBA_Capability_Map.xlsx into IRElements."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['capability_map_with_description']

    elements = []
    for r in range(3, ws.max_row + 1):
        l1 = ws.cell(r, 1).value
        l2 = ws.cell(r, 2).value
        l3 = ws.cell(r, 3).value
        core_obj = ws.cell(r, 4).value
        desc = ws.cell(r, 5).value

        if not l3:
            continue

        name = normalise(l3)
        el = IRElement(
            ir_element_id=make_id("cap", name),
            element_type=ElementType.CAPABILITY,
            name=name,
            description=normalise(desc) if desc else None,
            discovered_by=DiscoveredBy.INGEST,
            confidence=Confidence(data_quality=0.95, interpretive_certainty=1.0, provenance_depth=1),
            source_refs=[source_id],
            resolution_status=ResolutionStatus.ACCEPTED,
            attributes={
                "level1_area": normalise(l1) if l1 else None,
                "level2_domain": normalise(l2) if l2 else None,
                "core_object": normalise(core_obj) if core_obj else None,
            },
        )
        elements.append(el)

    return elements


# ── Value Stream Parser ───────────────────────────────────────────

def parse_value_streams(filepath: str, source_id: str, cap_elements: list[IRElement]) -> IREngagementModel:
    """Parse IIBA_Value_Streams_.xlsx into a full IREngagementModel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Build capability lookup for matching
    cap_lookup: dict[str, IRElement] = {}
    for cap in cap_elements:
        cap_lookup[normalise(cap.name).lower()] = cap

    ir = IREngagementModel(engagement_id="iiba_engagement_v1")

    # Register sources
    ir.sources.append(SourceRecord(
        source_id="src_vs_spreadsheet",
        document_type="spreadsheet",
        authority_level="formal",
        filename="IIBA_Value_Streams_.xlsx",
        ingestion_method="parser",
    ))
    ir.sources.append(SourceRecord(
        source_id="src_capability_map",
        document_type="capability_map",
        authority_level="formal",
        filename="IIBA_Capability_Map.xlsx",
        ingestion_method="parser",
    ))

    # Add capability elements
    for cap in cap_elements:
        ir.add_element(cap)

    # Parse each VS sheet
    vs_sheets = [s for s in wb.sheetnames if s[0].isdigit()]

    for sheet_name in vs_sheets:
        ws = wb[sheet_name]
        _parse_vs_sheet(ws, ir, source_id, cap_lookup)

    return ir


def _parse_vs_sheet(ws, ir: IREngagementModel, source_id: str, cap_lookup: dict):
    """Parse a single VS sheet into IR elements."""

    # VS name from row 1
    vs_name = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            vs_name = normalise(v)
            # Strip leading number: "1. Member Engagement..." → "Member Engagement..."
            vs_name = re.sub(r'^\d+\.\s*', '', vs_name)
            break

    if not vs_name:
        return

    # Purpose from row 3
    purpose = normalise(ws.cell(3, 3).value or "")

    # Find stage columns (row 6, columns with content)
    stage_cols = []
    for c in range(3, ws.max_column + 1):
        if ws.cell(6, c).value:
            stage_cols.append(c)

    if not stage_cols:
        return

    vs_id = make_id("vs", vs_name)

    # Create VS element
    activity_ids = []
    outcome_ids = []

    for idx, col in enumerate(stage_cols):
        stage_desc = normalise(ws.cell(6, col).value or "")
        # Use hand-crafted stage name if available, else derive
        stage_name = STAGE_NAMES.get((vs_name, idx), derive_stage_name(stage_desc))
        act_id = make_id("act", f"{vs_name}_{stage_name}")
        activity_ids.append(act_id)

        # Entry criteria (row 7)
        entry = normalise(ws.cell(7, col).value or "")
        # Exit criteria (row 8)
        exit_cr = normalise(ws.cell(8, col).value or "")
        # Value item / outcome (row 9)
        value_item = normalise(ws.cell(9, col).value or "")

        # Create outcome elements for entry and exit
        pre_outcome_id = None
        post_outcome_id = None

        if entry:
            pre_name = f"{stage_name} — Entry: {entry[:80]}"
            pre_outcome_id = make_id("outcome", f"{vs_name}_{stage_name}_entry")
            ir.add_element(IRElement(
                ir_element_id=pre_outcome_id,
                element_type=ElementType.OUTCOME,
                name=pre_name,
                description=entry,
                source_refs=[source_id],
                confidence=Confidence(data_quality=0.85, interpretive_certainty=0.85),
            ))
            outcome_ids.append(pre_outcome_id)

        if exit_cr or value_item:
            post_desc = f"Exit: {exit_cr}" + (f" | Value: {value_item}" if value_item else "")
            post_name = f"{stage_name} — Exit: {(exit_cr or value_item)[:80]}"
            post_outcome_id = make_id("outcome", f"{vs_name}_{stage_name}_exit")
            ir.add_element(IRElement(
                ir_element_id=post_outcome_id,
                element_type=ElementType.OUTCOME,
                name=post_name,
                description=post_desc,
                source_refs=[source_id],
                confidence=Confidence(data_quality=0.85, interpretive_certainty=0.85),
            ))
            outcome_ids.append(post_outcome_id)

        # Stakeholders / Roles (rows 10-13)
        role_ids = []
        for r in range(10, 14):
            role_name = ws.cell(r, col).value
            if role_name:
                role_name = normalise(role_name)
                role_id = make_id("role", role_name)
                # Deduplicate
                existing = ir.find_by_name(ElementType.ROLE, role_name)
                if not existing:
                    ir.add_element(IRElement(
                        ir_element_id=role_id,
                        element_type=ElementType.ROLE,
                        name=role_name,
                        source_refs=[source_id],
                        confidence=Confidence(data_quality=0.85, interpretive_certainty=0.9),
                    ))
                role_ids.append(role_id)

        # Metrics (rows 14-16)
        metric_ids = []
        for r in range(14, 17):
            metric_name = ws.cell(r, col).value
            if metric_name:
                # Some cells have multiple metrics separated by semicolons or newlines
                for m in re.split(r'[;\n]', metric_name):
                    m = normalise(m)
                    if not m:
                        continue
                    metric_id = make_id("metric", m)
                    existing = ir.find_by_name(ElementType.METRIC, m)
                    if not existing:
                        ir.add_element(IRElement(
                            ir_element_id=metric_id,
                            element_type=ElementType.METRIC,
                            name=m,
                            source_refs=[source_id],
                            confidence=Confidence(data_quality=0.8, interpretive_certainty=0.9),
                        ))
                    metric_ids.append(metric_id)

        # Capabilities (rows 18-22)
        cap_ids = []
        for r in range(18, 23):
            cap_name = ws.cell(r, col).value
            if cap_name:
                cap_name_norm = normalise(cap_name)
                # Try to match against capability map
                cap_key = cap_name_norm.lower()
                matched_cap = cap_lookup.get(cap_key)
                if matched_cap:
                    cap_ids.append(matched_cap.ir_element_id)
                else:
                    # Fuzzy match: try stripping trailing whitespace variations
                    found = False
                    for k, v in cap_lookup.items():
                        if k.replace(' ', '') == cap_key.replace(' ', ''):
                            cap_ids.append(v.ir_element_id)
                            found = True
                            break
                    if not found:
                        # Create new capability (not in cap map — flag as gap)
                        new_cap_id = make_id("cap", cap_name_norm)
                        existing = ir.find_by_name(ElementType.CAPABILITY, cap_name_norm)
                        if not existing:
                            ir.add_element(IRElement(
                                ir_element_id=new_cap_id,
                                element_type=ElementType.CAPABILITY,
                                name=cap_name_norm,
                                description=f"Capability referenced in VS '{vs_name}' but not found in capability map",
                                source_refs=[source_id],
                                confidence=Confidence(data_quality=0.7, interpretive_certainty=0.7),
                                conflict_flags=[],
                                resolution_status=ResolutionStatus.ACCEPTED,
                            ))
                        cap_ids.append(new_cap_id)

        # Create Activity element
        ir.add_element(IRElement(
            ir_element_id=act_id,
            element_type=ElementType.ACTIVITY,
            name=stage_name,
            description=stage_desc,
            source_refs=[source_id],
            confidence=Confidence(data_quality=0.9, interpretive_certainty=0.85),
            attributes={
                "vs_id": vs_id,
                "stage_index": idx,
                "pre_outcome_id": pre_outcome_id,
                "post_outcome_id": post_outcome_id,
                "role_ids": role_ids,
                "metric_ids": metric_ids,
                "capability_ids": cap_ids,
                "entry_criteria": entry,
                "exit_criteria": exit_cr,
                "value_item": value_item,
            },
        ))

    # Create VS element
    ir.add_element(IRElement(
        ir_element_id=vs_id,
        element_type=ElementType.VALUE_STREAM,
        name=vs_name,
        description=purpose,
        source_refs=[source_id],
        confidence=Confidence(data_quality=0.95, interpretive_certainty=1.0),
        attributes={
            "activity_ids": activity_ids,
            "outcome_ids": outcome_ids,
            "stage_count": len(stage_cols),
        },
    ))


# ── Main ──────────────────────────────────────────────────────────

def main():
    import json

    fixtures = Path(__file__).parent.parent / "fixtures"
    outputs = Path(__file__).parent.parent / "outputs"
    outputs.mkdir(exist_ok=True)

    # Step 1: Parse capability map
    print("Parsing capability map...")
    caps = parse_capability_map(
        str(fixtures / "IIBA_Capability_Map.xlsx"),
        source_id="src_capability_map",
    )
    print(f"  Capabilities from map: {len(caps)}")

    # Step 2: Parse value streams (with cap matching)
    print("\nParsing value streams...")
    ir = parse_value_streams(
        str(fixtures / "IIBA_Value_Streams_.xlsx"),
        source_id="src_vs_spreadsheet",
        cap_elements=caps,
    )

    # Summary
    summary = ir.summary()
    print(f"\n{'='*50}")
    print("IR Summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
    print(f"  sources: {len(ir.sources)}")

    # Check for unmatched capabilities
    vs_cap_ids = set()
    for act in ir.elements["activities"]:
        for cid in act.attributes.get("capability_ids", []):
            vs_cap_ids.add(cid)

    map_cap_ids = set(c.ir_element_id for c in caps)
    unmatched = vs_cap_ids - map_cap_ids
    matched = vs_cap_ids & map_cap_ids
    print(f"\n  Capability matching:")
    print(f"    Matched to cap map: {len(matched)}")
    print(f"    Not in cap map (new): {len(unmatched)}")
    if unmatched:
        for uid in sorted(unmatched):
            el = next((e for e in ir.elements["capabilities"] if e.ir_element_id == uid), None)
            if el:
                print(f"      - {el.name}")

    # List VS and their stages
    print(f"\nValue Streams:")
    for vs in ir.elements["value_streams"]:
        acts = [a for a in ir.elements["activities"] if a.attributes.get("vs_id") == vs.ir_element_id]
        print(f"\n  {vs.name} ({len(acts)} stages):")
        for a in sorted(acts, key=lambda x: x.attributes.get("stage_index", 0)):
            caps_count = len(a.attributes.get("capability_ids", []))
            roles_count = len(a.attributes.get("role_ids", []))
            metrics_count = len(a.attributes.get("metric_ids", []))
            print(f"    {a.attributes['stage_index']+1}. {a.name}")
            print(f"       caps={caps_count} roles={roles_count} metrics={metrics_count}")

    # Export IR as JSON
    ir_path = outputs / "iiba_ir.json"
    with open(ir_path, 'w') as f:
        json.dump(ir.to_dict(), f, indent=2)
    print(f"\nIR written to: {ir_path}")

    return ir


if __name__ == "__main__":
    main()
